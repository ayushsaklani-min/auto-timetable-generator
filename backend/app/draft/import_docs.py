"""Import subject/faculty documents into the Step 2 course paste format."""
from __future__ import annotations

import csv
import io
import json
import os
import re
import zipfile
from dataclasses import dataclass, field
from functools import lru_cache
from typing import Iterable, Optional
from xml.etree import ElementTree as ET

import httpx

from ..models.domain import CourseType
from .from_paste import ParsedCourse, parse_courses_text

_DOCX_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
_COURSE_CODE_RE = re.compile(r"\b([A-Z]{2,}[A-Z0-9_/-]*\d[A-Z0-9_/-]*)\b")
_HEADER_LINES = [
    "# CODE, NAME, CREDITS, TYPE, FACULTY1; FACULTY2; ...",
    "# TYPE: theory | lab | lab pair=CODE | activity | activity locked=DAY:slots | combined",
    "# Faculty: list one per section in order, or 'auto', 'xN', or 'same-as=CODE'",
    "",
]
_HEADER_ALIASES = {
    "code": ["course code", "subject code", "paper code", "code"],
    "name": ["course name", "subject name", "subject", "course", "title", "paper"],
    "credits": ["credits", "credit", "credits/hrs", "credit hrs", "hrs", "hours"],
    "type": ["type", "category", "mode", "kind"],
    "faculty": ["faculty", "faculties", "teacher", "teachers", "staff", "instructor", "handled by"],
    "sections": ["section(s)", "sections", "section", "sec"],
}
_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "tif", "tiff", "gif"}
_SPREADSHEET_EXTENSIONS = {"xlsx", "xlsm", "csv"}
_GOOGLE_VISION_PAGE_CHUNK = 5
_OPTIIC_API_URL = "https://api.optiic.dev/process"
_COURSE_CATEGORY_PREFIXES = (
    "PCC",
    "PCCL",
    "IPCC",
    "ESC",
    "AEC",
    "SEC",
    "BSC",
    "PLC",
    "ETC",
    "UHV",
    "NCMC",
    "MC",
    "HSMC",
)
_TIMETABLE_NOISE_TERMS = (
    "class room",
    "with effect",
    "version",
    "time table",
    "tea",
    "lunch",
    "break",
    "monday",
    "tuesday",
    "wednesday",
    "thursday",
    "friday",
    "saturday",
    "note: no classes",
    "chief time table officer",
    "dean",
    "principal",
)
_DAY_OR_BREAK_TERMS = ("tea", "lunch", "break", "monday", "tuesday", "wednesday", "thursday", "friday", "saturday")
_TITLE_RE = re.compile(r"\b(?:dr|prof|professor)\.?\b", re.IGNORECASE)
_COURSE_CODE_SPACED_RE = re.compile(r"\b([A-Z]{2,}\s+\d{2,3}[A-Z]?)\b")


@dataclass
class DraftCourseRow:
    code: str
    name: str
    credits: int
    kind: str
    pair_course: str = ""
    locked_day: str = ""
    locked_slots: list[int] = field(default_factory=list)
    combined: bool = False
    faculty_directive: str = ""
    faculty_names: list[str] = field(default_factory=list)


@dataclass
class ImportedRow:
    code: str = ""
    name: str = ""
    credits: Optional[int] = None
    kind: str = ""
    pair_course: str = ""
    locked_day: str = ""
    locked_slots: list[int] = field(default_factory=list)
    combined: bool = False
    faculty_directive: str = ""
    faculty_names: list[str] = field(default_factory=list)


@dataclass
class SpreadsheetCourseRecord:
    code: str = ""
    name: str = ""
    credits: Optional[int] = None
    kind: str = ""
    pair_course: str = ""
    locked_day: str = ""
    locked_slots: list[int] = field(default_factory=list)
    combined: bool = False
    faculty_directive: str = ""
    faculty_pool: list[str] = field(default_factory=list)
    section_faculty: dict[str, list[str]] = field(default_factory=dict)


@dataclass
class DocumentImportSummary:
    filename: str
    format: str
    course_rows: int = 0
    faculty_rows: int = 0
    warnings: list[str] = field(default_factory=list)


def import_course_documents(
    files: list[tuple[str, bytes]],
    section_ids: list[str],
    existing_text: str = "",
    replace_existing: bool = True,
) -> dict:
    base_rows = [] if replace_existing else _parse_existing_rows(existing_text)
    course_index = {row.code.upper(): row for row in base_rows if row.code.strip()}
    name_index = {_course_name_key(row.name): row for row in base_rows if row.name.strip()}

    documents: list[DocumentImportSummary] = []
    warnings: list[str] = []
    imported_rows = 0
    imported_faculty = 0

    def upsert(candidate: ImportedRow) -> None:
        nonlocal imported_rows, imported_faculty
        if not candidate.code.strip() and not candidate.name.strip():
            return
        target: Optional[DraftCourseRow] = None
        code_key = candidate.code.upper().strip()
        if code_key:
            target = course_index.get(code_key)
        if target is None and candidate.name.strip():
            target = name_index.get(_course_name_key(candidate.name))

        if target is None:
            target = DraftCourseRow(
                code=(candidate.code or _slugify(candidate.name) or f"COURSE_{len(base_rows) + 1}").upper(),
                name=candidate.name.strip() or candidate.code.strip() or f"Course {len(base_rows) + 1}",
                credits=candidate.credits or _default_credits(candidate.kind, candidate.name, candidate.code),
                kind=candidate.kind or _infer_kind(candidate.name, candidate.code),
            )
            base_rows.append(target)
            course_index[target.code.upper()] = target
            name_index[_course_name_key(target.name)] = target
            imported_rows += 1

        current_name_quality = _name_quality(target.name)
        candidate_name = _clean_candidate_name(candidate.name, candidate.code)
        candidate_name_quality = _name_quality(candidate_name)
        if candidate_name and candidate_name_quality >= current_name_quality:
            target.name = candidate_name
            name_index[_course_name_key(target.name)] = target
        if candidate.credits is not None:
            if (
                target.credits <= 0
                or target.credits == _default_credits(target.kind, target.name, target.code)
                or candidate_name_quality >= current_name_quality
            ):
                target.credits = max(0, int(candidate.credits))
        if candidate.kind and candidate_name_quality >= current_name_quality:
            target.kind = candidate.kind
        if candidate.pair_course:
            target.pair_course = candidate.pair_course
        if candidate.locked_day:
            target.locked_day = candidate.locked_day
        if candidate.locked_slots:
            target.locked_slots = list(candidate.locked_slots)
        if candidate.combined:
            target.combined = True
        if candidate.faculty_directive:
            target.faculty_directive = candidate.faculty_directive
            target.faculty_names = []
        if candidate.faculty_names:
            target.faculty_names = _merge_faculty_names(target.faculty_names, candidate.faculty_names)
            target.faculty_directive = ""
            imported_faculty += 1

    for filename, data in files:
        summary = _parse_document(filename, data, section_ids)
        documents.append(summary["summary"])
        warnings.extend(summary["summary"].warnings)
        for row in summary["rows"]:
            upsert(row)

    if not base_rows:
        detail = next(
            (
                warning
                for warning in warnings
                if "Google Vision" in warning
                or "Optiic" in warning
                or "Unsupported file type" in warning
            ),
            "",
        )
        if detail:
            raise ValueError(
                f"No subject or faculty rows could be recognised from the uploaded files. {detail}"
            )
        raise ValueError("No subject or faculty rows could be recognised from the uploaded files.")

    raw_text = _serialize_rows(base_rows)
    return {
        "raw_text": raw_text,
        "summary": {
            "documents": len(documents),
            "courses": len(base_rows),
            "imported_courses": imported_rows,
            "faculty_mappings": imported_faculty,
        },
        "documents_info": [
            {
                "filename": doc.filename,
                "format": doc.format,
                "course_rows": doc.course_rows,
                "faculty_rows": doc.faculty_rows,
                "warnings": doc.warnings,
            }
            for doc in documents
        ],
        "warnings": warnings,
    }


def _parse_existing_rows(text: str) -> list[DraftCourseRow]:
    try:
        parsed = parse_courses_text(text)
    except Exception:
        return []
    return [_parsed_to_draft(row) for row in parsed]


def _parsed_to_draft(row: ParsedCourse) -> DraftCourseRow:
    faculty_directive = row.faculty_spec.strip()
    faculty_names: list[str] = []
    if faculty_directive and not _is_faculty_directive(faculty_directive):
        faculty_names = _split_faculty_names(faculty_directive)
        faculty_directive = ""
    return DraftCourseRow(
        code=row.code,
        name=row.name,
        credits=row.credits,
        kind=_kind_label(row.type),
        pair_course=row.pair_course or "",
        locked_day=row.locked_day or "",
        locked_slots=list(row.locked_slots or []),
        combined=row.is_combined,
        faculty_directive=faculty_directive,
        faculty_names=faculty_names,
    )


def _kind_label(kind: CourseType) -> str:
    if kind is CourseType.LAB:
        return "lab"
    if kind is CourseType.ACTIVITY:
        return "activity"
    return "theory"


def _parse_document(filename: str, data: bytes, section_ids: list[str]) -> dict:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    summary = DocumentImportSummary(filename=filename, format=ext or "unknown")
    rows: list[ImportedRow] = []

    if ext == "pdf":
        tables, lines, extract_warnings = _extract_pdf(data, filename)
    elif ext == "docx":
        tables, lines, extract_warnings = _extract_docx(data)
    elif ext in _SPREADSHEET_EXTENSIONS:
        tables, lines, extract_warnings = _extract_spreadsheet(data, filename, section_ids)
    elif ext in _IMAGE_EXTENSIONS:
        tables, lines, extract_warnings = _extract_image(data, filename)
    else:
        raise ValueError(
            f"Unsupported file type for {filename!r}. Upload PDF, DOCX, XLSX, CSV, PNG, JPG, TIFF, or GIF files."
        )

    summary.warnings.extend(extract_warnings)

    seen = set()
    if ext in _SPREADSHEET_EXTENSIONS:
        for row in _parse_spreadsheet_tables(tables, section_ids):
            key = (row.code.upper(), row.name.upper(), tuple(row.faculty_names), row.faculty_directive)
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
            if row.code or row.name:
                summary.course_rows += 1
            if row.faculty_names or row.faculty_directive:
                summary.faculty_rows += 1

        if not rows:
            summary.warnings.append(
                "No structured spreadsheet rows were recognised. Include course/faculty/section headers for best results."
            )
        return {"summary": summary, "rows": rows}

    for table in tables:
        for row in _parse_table_rows(table, section_ids):
            key = (row.code.upper(), row.name.upper(), tuple(row.faculty_names))
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
            if row.code or row.name:
                summary.course_rows += 1
            if row.faculty_names:
                summary.faculty_rows += 1

    line_tables = _collect_line_tables(lines)
    for table in line_tables:
        for row in _parse_table_rows(table, section_ids):
            key = (row.code.upper(), row.name.upper(), tuple(row.faculty_names))
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
            if row.code or row.name:
                summary.course_rows += 1
            if row.faculty_names:
                summary.faculty_rows += 1

    catalog_rows = _parse_course_catalog_blocks(lines)
    for row in catalog_rows:
        key = (row.code.upper(), row.name.upper(), tuple(row.faculty_names))
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)
        if row.code or row.name:
            summary.course_rows += 1
        if row.faculty_names:
            summary.faculty_rows += 1

    free_text_hits = _parse_free_text(lines)
    for row in free_text_hits:
        key = (row.code.upper(), row.name.upper(), tuple(row.faculty_names))
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)
        if row.code or row.name:
            summary.course_rows += 1
        if row.faculty_names:
            summary.faculty_rows += 1

    if not rows:
        summary.warnings.append(
            "No structured rows were recognised. Use documents with course/faculty tables for best results."
        )

    return {"summary": summary, "rows": rows}


def _extract_pdf(data: bytes, filename: str) -> tuple[list[list[list[str]]], list[str], list[str]]:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    lines: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        lines.extend(_clean_lines(text.splitlines()))
    warnings: list[str] = []

    if _should_run_google_ocr(lines, page_count=len(reader.pages)):
        ocr_lines, ocr_warning = _ocr_pdf_with_provider(
            data,
            page_count=len(reader.pages),
            filename=filename,
        )
        if ocr_lines:
            lines = ocr_lines
            warnings.append(f"{filename}: used OCR fallback for a scanned or low-text PDF.")
        elif ocr_warning:
            warnings.append(f"{filename}: {ocr_warning}")

    return [], lines, warnings


def _extract_docx(data: bytes) -> tuple[list[list[list[str]]], list[str], list[str]]:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)

    tables: list[list[list[str]]] = []
    for tbl in root.findall(".//w:tbl", _DOCX_NS):
        parsed_rows: list[list[str]] = []
        for tr in tbl.findall("./w:tr", _DOCX_NS):
            row: list[str] = []
            for tc in tr.findall("./w:tc", _DOCX_NS):
                text = " ".join(t.text or "" for t in tc.findall(".//w:t", _DOCX_NS)).strip()
                row.append(_collapse_ws(text))
            if any(cell for cell in row):
                parsed_rows.append(row)
        if len(parsed_rows) >= 2:
            tables.append(parsed_rows)

    lines = []
    for para in root.findall(".//w:p", _DOCX_NS):
        text = " ".join(t.text or "" for t in para.findall(".//w:t", _DOCX_NS)).strip()
        if text:
            lines.append(_collapse_ws(text))
    return tables, _clean_lines(lines), []


def _extract_spreadsheet(
    data: bytes,
    filename: str,
    section_ids: list[str],
) -> tuple[list[list[list[str]]], list[str], list[str]]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext == "csv":
        text = data.decode("utf-8-sig", errors="replace")
        rows = [[_cell_to_text(cell) for cell in row] for row in csv.reader(io.StringIO(text))]
        return _detect_spreadsheet_tables([rows], section_ids), [], []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], [], ["Excel import needs `openpyxl` to be installed."]

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    matrices: list[list[list[str]]] = []
    try:
        for sheet in workbook.worksheets:
            rows: list[list[str]] = []
            for values in sheet.iter_rows(values_only=True):
                cells = [_cell_to_text(value) for value in values]
                while cells and not cells[-1]:
                    cells.pop()
                if any(cells):
                    rows.append(cells)
                elif rows:
                    rows.append([])
            matrices.append(rows)
    finally:
        workbook.close()
    return _detect_spreadsheet_tables(matrices, section_ids), [], []


def _extract_image(data: bytes, filename: str) -> tuple[list[list[list[str]]], list[str], list[str]]:
    lines, ocr_warning = _ocr_image_bytes(data, filename=filename, mime_type=_guess_mime_type(filename))
    warnings: list[str] = []
    if lines:
        warnings.append(f"{filename}: used OCR for the uploaded image.")
    elif ocr_warning:
        warnings.append(f"{filename}: {ocr_warning}")
    else:
        warnings.append(f"{filename}: no text was recognised from the uploaded image.")
    return [], lines, warnings


def _detect_spreadsheet_tables(
    matrices: list[list[list[str]]],
    section_ids: list[str],
) -> list[list[list[str]]]:
    tables: list[list[list[str]]] = []
    for rows in matrices:
        index = 0
        while index < len(rows):
            row = rows[index]
            if not row or not _looks_like_table_header(row, section_ids):
                index += 1
                continue

            table: list[list[str]] = [row]
            index += 1
            blank_count = 0
            while index < len(rows):
                current = rows[index]
                if current and _looks_like_table_header(current, section_ids):
                    break
                if not current:
                    blank_count += 1
                    if blank_count >= 2:
                        break
                    index += 1
                    continue
                blank_count = 0
                table.append(current)
                index += 1

            if len(table) >= 2:
                tables.append(table)
    return tables


def _parse_spreadsheet_tables(tables: list[list[list[str]]], section_ids: list[str]) -> list[ImportedRow]:
    records: dict[str, SpreadsheetCourseRecord] = {}
    section_lookup = {sid.upper(): sid for sid in section_ids}

    for table in tables:
        if len(table) < 2:
            continue
        header_map = _map_headers(table[0], section_ids)
        data_rows = table[1:]
        if header_map.get("sections") is not None and header_map["faculty"]:
            _merge_faculty_assignment_table(records, header_map, data_rows, section_ids, section_lookup)
            continue
        if header_map["section_faculty"]:
            _merge_section_matrix_table(records, header_map, data_rows, section_ids)
            continue
        _merge_course_catalog_table(records, header_map, data_rows)

    _apply_4th_sem_defaults(records)
    return [
        row
        for row in (_record_to_imported_row(record, section_ids) for record in records.values())
        if row and not _is_total_like(row.code, row.name)
    ]


def _record_for(
    records: dict[str, SpreadsheetCourseRecord],
    *,
    code: str,
    name: str = "",
) -> SpreadsheetCourseRecord:
    code_key = code.upper().strip()
    if not code_key and name:
        code_key = _slugify(name)
    if not code_key:
        code_key = f"COURSE_{len(records) + 1}"
    record = records.get(code_key)
    if record is None:
        record = SpreadsheetCourseRecord(code=code_key)
        records[code_key] = record
    if name and _name_quality(name) >= _name_quality(record.name):
        record.name = _clean_candidate_name(name, code_key)
    return record


def _merge_course_catalog_table(
    records: dict[str, SpreadsheetCourseRecord],
    header_map: dict,
    rows: list[list[str]],
) -> None:
    for row in rows:
        code = _value_at(row, header_map.get("code"))
        name = _value_at(row, header_map.get("name"))
        code = _extract_structured_code(code)
        if _is_total_like(code, name):
            continue
        if not code and not name:
            continue

        record = _record_for(records, code=code, name=name)
        if header_map.get("credits") is not None:
            credit_text = _value_at(row, header_map["credits"])
            credits = _parse_credit_value(credit_text)
            if credits is None:
                credits = _parse_int(credit_text)
            if credits is not None:
                record.credits = credits
        if header_map.get("type") is not None:
            _apply_type_to_record(record, _value_at(row, header_map["type"]))
        if not record.kind:
            record.kind = _infer_kind(record.name, record.code)


def _merge_section_matrix_table(
    records: dict[str, SpreadsheetCourseRecord],
    header_map: dict,
    rows: list[list[str]],
    section_ids: list[str],
) -> None:
    for row in rows:
        code = _extract_structured_code(_value_at(row, header_map.get("code")))
        name = _value_at(row, header_map.get("name"))
        if _is_total_like(code, name) or (not code and not name):
            continue
        record = _record_for(records, code=code, name=name)
        if header_map.get("credits") is not None:
            credits = _parse_int(_value_at(row, header_map["credits"]))
            if credits is not None:
                record.credits = credits
        if header_map.get("type") is not None:
            _apply_type_to_record(record, _value_at(row, header_map["type"]))
        for section_id, idx in header_map["section_faculty"]:
            names = _split_faculty_names(_value_at(row, idx))
            if names:
                record.section_faculty[section_id] = _merge_faculty_names(
                    record.section_faculty.get(section_id, []),
                    names,
                )
        if not header_map["section_faculty"]:
            for name in _collect_generic_faculty(header_map, row):
                record.faculty_pool = _merge_faculty_names(record.faculty_pool, [name])
        if not record.kind:
            record.kind = _infer_kind(record.name, record.code)


def _merge_faculty_assignment_table(
    records: dict[str, SpreadsheetCourseRecord],
    header_map: dict,
    rows: list[list[str]],
    section_ids: list[str],
    section_lookup: dict[str, str],
) -> None:
    faculty_cols = header_map["faculty"]
    faculty_idx = faculty_cols[0] if faculty_cols else None
    section_idx = header_map.get("sections")
    for row in rows:
        code = _extract_structured_code(_value_at(row, header_map.get("code")))
        name = _value_at(row, header_map.get("name"))
        if _is_total_like(code, name) or (not code and not name):
            continue
        faculty_names = _split_faculty_names(_value_at(row, faculty_idx))
        if not faculty_names:
            continue

        record = _record_for(records, code=code, name=name)
        if header_map.get("credits") is not None:
            credits = _parse_int(_value_at(row, header_map["credits"]))
            if credits is not None:
                record.credits = credits
        if header_map.get("type") is not None:
            _apply_type_to_record(record, _value_at(row, header_map["type"]))

        row_sections = _parse_section_targets(_value_at(row, section_idx), section_ids, section_lookup)
        if row_sections:
            for section_id in row_sections:
                record.section_faculty[section_id] = _merge_faculty_names(
                    record.section_faculty.get(section_id, []),
                    faculty_names,
                )
        else:
            record.faculty_pool = _merge_faculty_names(record.faculty_pool, faculty_names)
        if not record.kind:
            record.kind = _infer_kind(record.name, record.code)


def _record_to_imported_row(record: SpreadsheetCourseRecord, section_ids: list[str]) -> ImportedRow | None:
    if not record.code and not record.name:
        return None
    kind = record.kind or _infer_kind(record.name, record.code)
    faculty_names = _faculty_for_record(record, section_ids)
    return ImportedRow(
        code=record.code,
        name=record.name or record.code,
        credits=record.credits if record.credits is not None else _default_credits(kind, record.name, record.code),
        kind=kind,
        pair_course=record.pair_course,
        locked_day=record.locked_day,
        locked_slots=list(record.locked_slots),
        combined=record.combined,
        faculty_directive=record.faculty_directive,
        faculty_names=faculty_names,
    )


def _faculty_for_record(record: SpreadsheetCourseRecord, section_ids: list[str]) -> list[str]:
    if record.faculty_pool and not record.section_faculty:
        return record.faculty_pool

    ordered_groups = [record.section_faculty.get(section_id, []) for section_id in section_ids]
    non_empty_groups = [group for group in ordered_groups if group]
    if len(non_empty_groups) == len(section_ids) and all(len(group) == 1 for group in non_empty_groups):
        return [group[0] for group in ordered_groups]

    # The current solver input has one faculty list per course. When the source
    # has multiple teachers per section, keep a de-duplicated pool rather than
    # forcing a fake section order.
    merged: list[str] = []
    for group in ordered_groups:
        merged = _merge_faculty_names(merged, group)
    merged = _merge_faculty_names(merged, record.faculty_pool)
    return merged


def _apply_4th_sem_defaults(records: dict[str, SpreadsheetCourseRecord]) -> None:
    for record in records.values():
        record.code = _normalize_code(record.code)
        record.name = _clean_candidate_name(record.name, record.code) or record.name
        if not record.kind:
            record.kind = _infer_kind(record.name, record.code)

    # The 4th-sem wizard already models BENGDIP2 through Saturday locked slots.
    records.pop("BENGDIP2", None)

    if "BAIL456B" in records and "BAIL456C" in records:
        records["BAIL456B"].kind = "lab"
        records["BAIL456C"].kind = "lab"
        records["BAIL456B"].pair_course = "BAIL456C"
        records["BAIL456C"].pair_course = "BAIL456B"

    if "BCSL404" in records:
        records["BCSL404"].kind = "lab"

    for code in ("BAI402", "BCS403"):
        record = records.get(code)
        if not record or record.credits != 4:
            continue
        record.credits = 3
        lab_code = f"{code}_LAB"
        lab = records.get(lab_code)
        if lab is None:
            lab = SpreadsheetCourseRecord(code=lab_code)
            records[lab_code] = lab
        lab.name = f"{record.name} Lab" if record.name else f"{code} Lab"
        lab.credits = 1
        lab.kind = "lab"
        if record.section_faculty:
            lab.section_faculty = {
                section_id: list(names)
                for section_id, names in record.section_faculty.items()
            }
        elif record.faculty_pool:
            lab.faculty_pool = list(record.faculty_pool)
        if code == "BAI402" and "BCSL404" in records:
            lab.pair_course = "BCSL404"
            records["BCSL404"].pair_course = lab_code

    if "NCMC" in records:
        records["NCMC"].kind = "activity"
    if "DEPT_ACT" in records:
        records["DEPT_ACT"].kind = "activity"


def _apply_type_to_record(record: SpreadsheetCourseRecord, value: str) -> None:
    kind, pair_course, locked_day, locked_slots, combined = _parse_type_details(
        value,
        record.name,
        record.code,
    )
    if kind:
        record.kind = kind
    if pair_course:
        record.pair_course = pair_course
    if locked_day:
        record.locked_day = locked_day
    if locked_slots:
        record.locked_slots = locked_slots
    if combined:
        record.combined = True


def _collect_generic_faculty(header_map: dict, row: list[str]) -> list[str]:
    faculty_names: list[str] = []
    for idx in header_map["faculty"]:
        faculty_names = _merge_faculty_names(faculty_names, _split_faculty_names(_value_at(row, idx)))
    return faculty_names


def _parse_section_targets(
    value: str,
    section_ids: list[str],
    section_lookup: dict[str, str],
) -> list[str]:
    normalized = _canonical_header(value)
    if not normalized:
        return []
    if "all" in normalized:
        return list(section_ids)
    found: list[str] = []
    for token in re.split(r"[^a-z0-9]+", normalized):
        section_id = section_lookup.get(token.upper())
        if section_id and section_id not in found:
            found.append(section_id)
    return found


def _value_at(row: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return _collapse_ws(row[index])


def _is_total_like(code: str, name: str) -> bool:
    text = f"{code} {name}".lower()
    return "total" in text or text.strip() in {"credits", "grand"}


def _should_run_google_ocr(lines: list[str], page_count: int) -> bool:
    if not lines:
        return True
    text = "\n".join(lines)
    alnum_count = sum(char.isalnum() for char in text)
    if alnum_count == 0:
        return True
    min_lines = max(3, page_count * 2)
    min_alnum = max(80, page_count * 60)
    return len(lines) < min_lines or alnum_count < min_alnum


@lru_cache(maxsize=1)
def _get_google_vision_client() -> tuple[object | None, str | None]:
    try:
        from google.cloud import vision
    except ImportError:
        return None, "Google Vision OCR is unavailable because `google-cloud-vision` is not installed."

    client_options = None
    api_endpoint = os.environ.get("GOOGLE_VISION_API_ENDPOINT", "").strip()
    if api_endpoint:
        client_options = {"api_endpoint": api_endpoint}

    raw_service_account = (
        os.environ.get("GOOGLE_APPLICATION_CREDENTIALS_JSON", "").strip()
        or os.environ.get("GOOGLE_VISION_SERVICE_ACCOUNT_JSON", "").strip()
    )
    try:
        if raw_service_account:
            from google.oauth2 import service_account

            info = json.loads(raw_service_account)
            credentials = service_account.Credentials.from_service_account_info(info)
            return (
                vision.ImageAnnotatorClient(
                    credentials=credentials,
                    client_options=client_options,
                ),
                None,
            )
        return vision.ImageAnnotatorClient(client_options=client_options), None
    except Exception as exc:
        return None, f"Google Vision OCR could not be initialized: {str(exc)[:200]}"


def _ocr_pdf_with_google_vision(data: bytes, page_count: int) -> tuple[list[str], str | None]:
    client, init_warning = _get_google_vision_client()
    if client is None:
        return [], init_warning

    from google.cloud import vision

    input_config = {"mime_type": "application/pdf", "content": data}
    features = [{"type_": vision.Feature.Type.DOCUMENT_TEXT_DETECTION}]
    lines: list[str] = []
    warnings: list[str] = []

    for start in range(1, page_count + 1, _GOOGLE_VISION_PAGE_CHUNK):
        pages = list(range(start, min(start + _GOOGLE_VISION_PAGE_CHUNK, page_count + 1)))
        response = client.batch_annotate_files(
            requests=[{"input_config": input_config, "features": features, "pages": pages}]
        )
        file_response = response.responses[0] if getattr(response, "responses", None) else None
        if file_response is None:
            warnings.append(
                f"Google Vision OCR returned no file response for PDF pages {pages[0]}-{pages[-1]}."
            )
            continue
        for image_response in file_response.responses:
            error_message = getattr(getattr(image_response, "error", None), "message", "")
            if error_message:
                warnings.append(
                    f"Google Vision OCR failed on PDF pages {pages[0]}-{pages[-1]}: {error_message}"
                )
                continue
            text = getattr(getattr(image_response, "full_text_annotation", None), "text", "") or ""
            lines.extend(_clean_lines(text.splitlines()))

    return lines, " ".join(warnings).strip() or None


def _ocr_image_with_google_vision(data: bytes) -> tuple[list[str], str | None]:
    client, init_warning = _get_google_vision_client()
    if client is None:
        return [], init_warning

    from google.cloud import vision

    image = vision.Image(content=data)
    response = client.document_text_detection(image=image)
    error_message = getattr(getattr(response, "error", None), "message", "")
    if error_message:
        return [], f"Google Vision OCR failed: {error_message}"
    text = getattr(getattr(response, "full_text_annotation", None), "text", "") or ""
    return _clean_lines(text.splitlines()), None


def _ocr_pdf_with_provider(data: bytes, page_count: int, filename: str) -> tuple[list[str], str | None]:
    api_key = os.environ.get("OPTIIC_API_KEY", "").strip()
    if api_key:
        return _ocr_pdf_with_optiic(data, filename)
    return _ocr_pdf_with_google_vision(data, page_count=page_count)


def _ocr_image_bytes(data: bytes, filename: str, mime_type: str) -> tuple[list[str], str | None]:
    api_key = os.environ.get("OPTIIC_API_KEY", "").strip()
    if api_key:
        return _ocr_image_with_optiic(data, filename=filename, mime_type=mime_type)
    return _ocr_image_with_google_vision(data)


def _ocr_pdf_with_optiic(data: bytes, filename: str) -> tuple[list[str], str | None]:
    lines, warning = _call_optiic(data, filename=filename, mime_type="application/pdf")
    if lines:
        return lines, None

    if warning and "pdf submissions" in warning.lower():
        page_images, render_warning = _render_pdf_pages_for_ocr(data, filename)
        if not page_images:
            return [], render_warning or warning

        merged: list[str] = []
        page_warnings: list[str] = []
        for page_name, page_bytes in page_images:
            page_lines, page_warning = _call_optiic(
                page_bytes,
                filename=page_name,
                mime_type="image/png",
            )
            if page_lines:
                merged.extend(page_lines)
            elif page_warning:
                page_warnings.append(f"{page_name}: {page_warning}")

        if merged:
            return (
                merged,
                "Used Optiic page-image fallback because the current Optiic plan rejected direct PDF OCR.",
            )
        if page_warnings:
            return [], " ".join(page_warnings)
        return [], warning

    return [], warning


def _ocr_image_with_optiic(data: bytes, filename: str, mime_type: str) -> tuple[list[str], str | None]:
    return _call_optiic(data, filename=filename, mime_type=mime_type)


def _call_optiic(data: bytes, filename: str, mime_type: str) -> tuple[list[str], str | None]:
    api_key = os.environ.get("OPTIIC_API_KEY", "").strip()
    if not api_key:
        return [], "Optiic OCR is unavailable because `OPTIIC_API_KEY` is not set."

    url = os.environ.get("OPTIIC_API_URL", _OPTIIC_API_URL).strip() or _OPTIIC_API_URL
    timeout = float(os.environ.get("OPTIIC_TIMEOUT_SEC", "120") or "120")

    try:
        with httpx.Client(timeout=timeout) as client:
            response = client.post(
                url,
                data={"apiKey": api_key},
                files={"image": (filename, data, mime_type)},
            )
    except Exception as exc:
        return [], f"Optiic OCR request failed: {str(exc)[:200]}"

    body = response.text.strip()
    if not response.is_success:
        detail = body[:300] or f"HTTP {response.status_code}"
        return [], f"Optiic OCR failed: {detail}"

    try:
        payload = response.json()
    except Exception:
        if body:
            return [], body[:300]
        return [], "Optiic OCR returned an empty response."

    text = (
        payload.get("text")
        or payload.get("fullText")
        or payload.get("full_text")
        or payload.get("data", {}).get("text")
        or ""
    )
    if not text and body and body[0] != "{":
        text = body
    lines = _clean_lines(text.splitlines())
    if lines:
        return lines, None
    return [], payload.get("error") or "Optiic OCR returned no text."


def _render_pdf_pages_for_ocr(data: bytes, filename: str) -> tuple[list[tuple[str, bytes]], str | None]:
    try:
        import fitz
    except ImportError:
        return [], "Optiic OCR PDF fallback needs `pymupdf` to render pages as images."

    try:
        doc = fitz.open(stream=data, filetype="pdf")
    except Exception as exc:
        return [], f"Unable to open {filename} for OCR rendering: {str(exc)[:200]}"

    images: list[tuple[str, bytes]] = []
    stem = _slugify(filename.rsplit(".", 1)[0]) or "page"
    try:
        for index, page in enumerate(doc):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            images.append((f"{stem}_page_{index + 1:02d}.png", pix.tobytes("png")))
    finally:
        doc.close()

    if not images:
        return [], f"{filename}: no pages were rendered for OCR."
    return images, None


def _guess_mime_type(filename: str) -> str:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    return {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "tif": "image/tiff",
        "tiff": "image/tiff",
        "gif": "image/gif",
        "bmp": "image/bmp",
        "pdf": "application/pdf",
    }.get(ext, "application/octet-stream")


def _collect_line_tables(lines: list[str]) -> list[list[list[str]]]:
    groups: list[list[list[str]]] = []
    current: list[list[str]] = []
    expected_cols = 0

    for line in lines:
        parts = _split_tabular_line(line)
        if parts:
            if not current:
                current = [parts]
                expected_cols = len(parts)
                continue
            if abs(len(parts) - expected_cols) <= 2:
                current.append(parts)
                expected_cols = max(expected_cols, len(parts))
                continue
            if len(current) >= 2:
                groups.append(current)
            current = [parts]
            expected_cols = len(parts)
            continue
        if len(current) >= 2:
            groups.append(current)
        current = []
        expected_cols = 0

    if len(current) >= 2:
        groups.append(current)
    return groups


def _parse_table_rows(rows: list[list[str]], section_ids: list[str]) -> list[ImportedRow]:
    if len(rows) < 2:
        return []

    header_map = _map_headers(rows[0], section_ids)
    data_rows = rows[1:]
    parsed = _parse_with_header(header_map, data_rows)
    if parsed:
        return parsed
    return _parse_without_header(rows, section_ids)


def _map_headers(header: list[str], section_ids: list[str]) -> dict:
    mapped = {
        "code": None,
        "name": None,
        "credits": None,
        "type": None,
        "sections": None,
        "faculty": [],
        "section_faculty": [],
    }
    section_lookup = {sid.strip().upper(): sid.strip().upper() for sid in section_ids if sid.strip()}

    for idx, raw in enumerate(header):
        cell = _canonical_header(raw)
        if not cell:
            continue
        section_id = _match_section_header(cell, section_lookup)
        if section_id:
            mapped["section_faculty"].append((section_id, idx))
            continue
        for key, aliases in _HEADER_ALIASES.items():
            if any(alias in cell for alias in aliases):
                if key == "faculty":
                    mapped["faculty"].append(idx)
                elif key == "sections":
                    if mapped[key] is None:
                        mapped[key] = idx
                elif mapped[key] is None:
                    mapped[key] = idx
                break
    return mapped


def _looks_like_table_header(row: list[str], section_ids: list[str]) -> bool:
    header_map = _map_headers(row, section_ids)
    has_course_identity = header_map["code"] is not None or header_map["name"] is not None
    has_load = header_map["credits"] is not None or header_map["type"] is not None
    has_faculty = bool(header_map["faculty"] or header_map["section_faculty"])
    has_mapping = has_faculty and header_map["sections"] is not None
    if has_course_identity and has_mapping:
        return True
    return has_course_identity and (has_load or has_faculty)


def _parse_with_header(header_map: dict, rows: list[list[str]]) -> list[ImportedRow]:
    if (
        header_map["code"] is None
        and header_map["name"] is None
        and not header_map["faculty"]
        and not header_map["section_faculty"]
    ):
        return []

    parsed: list[ImportedRow] = []
    for row in rows:
        candidate = ImportedRow()
        if header_map["code"] is not None and header_map["code"] < len(row):
            candidate.code = _extract_structured_code(row[header_map["code"]])
        if header_map["name"] is not None and header_map["name"] < len(row):
            candidate.name = row[header_map["name"]].strip()
        if header_map["credits"] is not None and header_map["credits"] < len(row):
            candidate.credits = _parse_int(row[header_map["credits"]])
        if header_map["type"] is not None and header_map["type"] < len(row):
            _apply_type_details(candidate, row[header_map["type"]])
        elif candidate.name or candidate.code:
            candidate.kind = _infer_kind(candidate.name, candidate.code)

        faculty_names: list[str] = []
        for idx in header_map["faculty"]:
            if idx < len(row):
                _apply_faculty_cell(candidate, row[idx], faculty_names)
        by_section: dict[str, list[str]] = {}
        for section_id, idx in header_map["section_faculty"]:
            if idx < len(row):
                names: list[str] = []
                _apply_faculty_cell(candidate, row[idx], names)
                if names:
                    by_section[section_id] = names
        for section_id in _section_order(header_map["section_faculty"]):
            faculty_names.extend(by_section.get(section_id, []))
        candidate.faculty_names = _merge_faculty_names([], faculty_names)

        if not candidate.code and candidate.name:
            candidate.code = _slugify(candidate.name).upper()

        if candidate.code or candidate.name or candidate.faculty_names:
            parsed.append(candidate)
    return parsed


def _parse_without_header(rows: list[list[str]], section_ids: list[str]) -> list[ImportedRow]:
    parsed: list[ImportedRow] = []
    if not rows:
        return parsed

    for row in rows:
        if not row:
            continue
        first = _extract_structured_code(row[0])
        if not first:
            continue
        candidate = ImportedRow(code=first)
        if len(row) >= 2:
            candidate.name = row[1].strip()
        if len(row) >= 3:
            maybe_credits = _parse_int(row[2])
            if maybe_credits is not None:
                candidate.credits = maybe_credits
                if len(row) >= 4:
                    candidate.faculty_names = _merge_faculty_names([], _split_faculty_names(row[3]))
            else:
                candidate.faculty_names = _merge_faculty_names([], _split_faculty_names(row[2]))
                if len(row) >= 4:
                    candidate.faculty_names = _merge_faculty_names(
                        candidate.faculty_names,
                        _split_faculty_names(row[3]),
                    )
        if len(row) >= 4:
            _apply_type_details(candidate, row[3])
        if not candidate.kind:
            candidate.kind = _infer_kind(candidate.name, candidate.code)
        parsed.append(candidate)
    return parsed


def _section_order(section_faculty_headers: list[tuple[str, int]]) -> list[str]:
    return [
        section_id
        for section_id, _ in sorted(section_faculty_headers, key=lambda item: item[1])
    ]


def _apply_faculty_cell(candidate: ImportedRow, value: str, out: list[str]) -> None:
    cleaned = _collapse_ws(value)
    if not cleaned:
        return
    if _is_faculty_directive(cleaned):
        candidate.faculty_directive = cleaned
        return
    out.extend(_split_faculty_names(cleaned))


def _apply_type_details(candidate: ImportedRow, value: str) -> None:
    kind, pair_course, locked_day, locked_slots, combined = _parse_type_details(
        value,
        candidate.name,
        candidate.code,
    )
    if kind:
        candidate.kind = kind
    if pair_course:
        candidate.pair_course = pair_course
    if locked_day:
        candidate.locked_day = locked_day
    if locked_slots:
        candidate.locked_slots = locked_slots
    if combined:
        candidate.combined = True


def _parse_free_text(lines: list[str]) -> list[ImportedRow]:
    parsed: list[ImportedRow] = []
    for line in lines:
        normalized = _normalize_ocr_text(line)
        if _should_skip_free_text_line(normalized):
            continue
        code = _extract_course_code(normalized)
        if not code:
            continue
        code, inline_name = _split_code_and_alias(code)
        if re.search(r"\b(dr|prof|professor|faculty|instructor|teacher)\b", normalized, re.IGNORECASE):
            faculty_text = normalized.replace(code, " ")
            if inline_name:
                faculty_text = faculty_text.replace(inline_name, " ")
            names = _split_faculty_names(faculty_text)
            if names:
                parsed.append(ImportedRow(code=code, faculty_names=names))
                continue
        credit_match = re.search(r"\b(\d{1,2})\b", normalized)
        title = normalized.replace(code, " ", 1)
        if inline_name:
            title = title.replace(inline_name, " ", 1)
        if credit_match:
            title = title.replace(credit_match.group(1), " ", 1)
        title = _clean_candidate_name(title.strip(" -:|"), code)
        if title and _name_quality(title) > 0:
            parsed.append(
                ImportedRow(
                    code=code,
                    name=title,
                    credits=int(credit_match.group(1)) if credit_match else None,
                    kind=_infer_kind(title, code),
                )
            )
    return parsed


def _parse_course_catalog_blocks(lines: list[str]) -> list[ImportedRow]:
    parsed: list[ImportedRow] = []
    in_catalog = False
    current: ImportedRow | None = None

    def flush() -> None:
        nonlocal current
        if current and current.code and (_name_quality(current.name) > 0 or current.faculty_names):
            if not current.kind:
                current.kind = _infer_kind(current.name, current.code)
            parsed.append(current)
        current = None

    for raw_line in lines:
        line = _normalize_ocr_text(raw_line)
        if not line:
            continue
        if _looks_like_catalog_header(line):
            in_catalog = True
            flush()
            continue
        if not in_catalog:
            continue
        if _looks_like_catalog_footer(line):
            flush()
            in_catalog = False
            continue
        if _looks_like_catalog_row_start(line):
            flush()
            current = _parse_catalog_row_start(line)
            continue
        if current is None:
            continue
        code = _extract_course_code(line)
        has_title = bool(_TITLE_RE.search(line))
        if code and not current.code:
            current.code = code
        credit = _parse_credit_value(line)
        if credit is not None and current.credits is None:
            current.credits = credit
        if has_title:
            faculty_names = _split_faculty_names(line)
            if faculty_names:
                current.faculty_names = _merge_faculty_names(current.faculty_names, faculty_names)
        working = line
        if code:
            working = re.sub(re.escape(code), " ", working, flags=re.IGNORECASE)
            base_code, alias = _split_code_and_alias(code)
            if base_code and base_code != code:
                working = re.sub(re.escape(base_code), " ", working, flags=re.IGNORECASE)
            if alias:
                working = re.sub(rf"\b{re.escape(alias)}\b", " ", working, flags=re.IGNORECASE)
        if credit is not None:
            working = re.sub(
                rf"(?<![A-Z0-9]){credit}(?:\s*\([^\)]*\))?",
                " ",
                working,
                count=1,
                flags=re.IGNORECASE,
            )
        cleaned = _clean_candidate_name(working, current.code or code)
        if cleaned and _name_quality(cleaned) > 0:
            current.name = _collapse_ws(f"{current.name} {cleaned}".strip())

    flush()
    return parsed


def _looks_like_catalog_header(line: str) -> bool:
    lowered = line.lower()
    return "course code" in lowered and "credits" in lowered and "faculty" in lowered


def _looks_like_catalog_footer(line: str) -> bool:
    lowered = line.lower()
    return (
        "time table officer" in lowered
        or lowered.startswith("note:")
        or lowered == "principal"
        or lowered.startswith("bms institute of technology")
        or lowered.startswith("department of ")
    )


def _looks_like_catalog_row_start(line: str) -> bool:
    stripped = line.strip()
    lowered = stripped.lower()
    if lowered.startswith("universal human values"):
        return True
    if lowered.startswith("english communication skill"):
        return True
    if lowered.startswith("ncmc-"):
        return True
    return any(stripped.upper().startswith(prefix) for prefix in _COURSE_CATEGORY_PREFIXES)


def _parse_catalog_row_start(line: str) -> ImportedRow:
    code = _extract_course_code(line)
    credit = _parse_credit_value(line)
    faculty_names = _split_faculty_names(line) if _TITLE_RE.search(line) else []
    cleaned = _clean_candidate_name(line, code)
    return ImportedRow(
        code=code,
        name=cleaned,
        credits=credit,
        kind=_infer_kind(line, code),
        faculty_names=faculty_names,
    )


def _serialize_rows(rows: Iterable[DraftCourseRow]) -> str:
    lines = list(_HEADER_LINES)
    for row in rows:
        kind_bits = [row.kind or _infer_kind(row.name, row.code)]
        if row.kind == "lab" and row.pair_course.strip():
            kind_bits.append(f"pair={row.pair_course.strip()}")
        if row.kind == "activity" and row.locked_day.strip() and row.locked_slots:
            slots = _format_slot_list(row.locked_slots)
            kind_bits.append(f"locked={row.locked_day.strip().upper()}:{slots}")
        if row.combined:
            kind_bits.append("combined")
        faculty_spec = row.faculty_directive.strip()
        if row.faculty_names:
            faculty_spec = "; ".join(row.faculty_names)
        elif not faculty_spec:
            faculty_spec = "auto"
        lines.append(
            f"{row.code.strip()}, {row.name.strip()}, {max(0, int(row.credits))}, {' '.join(kind_bits)}, {faculty_spec}"
        )
    return "\n".join(lines)


def _split_tabular_line(line: str) -> Optional[list[str]]:
    raw = line.strip()
    if not raw:
        return None
    for pattern in (r"\s*\|\s*", r"\t+", r" {2,}"):
        parts = [_collapse_ws(part) for part in re.split(pattern, raw) if _collapse_ws(part)]
        if len(parts) >= 2:
            return parts
    return None


def _split_faculty_names(value: str) -> list[str]:
    if not value:
        return []
    cleaned = _normalize_ocr_text(value)
    cleaned = re.sub(r"\b(and|&)\b", ",", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bI(?=\s+(?:Dr|Prof)\.?)", ";", cleaned)
    titled = re.findall(
        r"(?:Dr|Prof|Professor)\.?\s*[A-Za-z][A-Za-z .'-]{1,80}",
        cleaned,
        flags=re.IGNORECASE,
    )
    if titled:
        return [
            _collapse_ws(re.sub(r"^[0-9]+", "", item).strip(" -:;,."))
            for item in titled
            if _collapse_ws(re.sub(r"^[0-9]+", "", item).strip(" -:;,."))
        ]
    parts = re.split(r"[;\n|,/]+", cleaned)
    return [
        _collapse_ws(re.sub(r"^[0-9]+", "", part).strip(" -:"))
        for part in parts
        if _collapse_ws(re.sub(r"^[0-9]+", "", part).strip(" -:"))
    ]


def _merge_faculty_names(current: list[str], incoming: list[str]) -> list[str]:
    seen = {_faculty_key(name) for name in current}
    out = list(current)
    for name in incoming:
        cleaned = _collapse_ws(name)
        if not cleaned:
            continue
        key = _faculty_key(cleaned)
        if key in seen:
            continue
        seen.add(key)
        out.append(cleaned)
    return out


def _extract_course_code(value: str) -> str:
    raw = (value or "").upper()
    match = _COURSE_CODE_RE.search(raw)
    if match:
        candidate = _normalize_code(match.group(1).strip())
        if _is_plausible_course_code(candidate):
            return candidate
    spaced = _COURSE_CODE_SPACED_RE.search(raw)
    if spaced:
        candidate = _normalize_code(spaced.group(1).replace(" ", ""))
        if _is_plausible_course_code(candidate):
            return candidate
    return ""


def _extract_structured_code(value: str) -> str:
    strict = _extract_course_code(value)
    if strict:
        return strict
    raw = _collapse_ws(value)
    if not raw:
        return ""
    code = re.sub(r"[^A-Za-z0-9_/-]+", "_", raw).strip("_").upper()
    if not code or len(code) > 48 or not any(char.isalpha() for char in code):
        return ""
    return code


def _parse_int(value: str) -> Optional[int]:
    match = re.search(r"\d+", value or "")
    return int(match.group(0)) if match else None


def _default_credits(kind: str, name: str, code: str) -> int:
    inferred = kind or _infer_kind(name, code)
    if inferred == "lab":
        return 1
    if inferred == "activity":
        return 0
    return 3


def _infer_kind(name: str, code: str) -> str:
    text = f"{name} {code}".lower()
    compact_code = re.sub(r"[^A-Z0-9]", "", (code or "").upper())
    if any(token in text for token in ("activity", "tutorial", "proctor", "remedial", "ncmc", "dept")):
        return "activity"
    if " lab" in text or text.endswith("_lab") or text.endswith("lab"):
        return "lab"
    if compact_code.startswith(("BCSL", "BAIL")):
        return "lab"
    return "theory"


def _parse_type_details(
    value: str,
    name: str = "",
    code: str = "",
) -> tuple[str, str, str, list[int], bool]:
    text = _collapse_ws(value)
    lowered = text.lower()
    kind = ""
    if re.search(r"\b(lab|laboratory|practical|pccl)\b", lowered):
        kind = "lab"
    if re.search(r"\b(activity|tutorial|proctor|remedial|ncmc|dept)\b", lowered):
        kind = "activity"
    if not kind and re.search(r"\b(theory|lecture|pcc|ipcc|esc|sec|aec|uhv|bsc|beng)\b", lowered):
        kind = "theory"
    if not kind:
        kind = _infer_kind(f"{name} {text}", code)

    pair_course = ""
    pair_match = re.search(r"\bpair\s*=\s*([A-Za-z0-9_/-]+)", text, flags=re.IGNORECASE)
    if pair_match:
        pair_course = _normalize_code(pair_match.group(1))

    locked_day = ""
    locked_slots: list[int] = []
    locked_match = re.search(
        r"\blocked\s*=\s*([A-Za-z]+)\s*:\s*([\d,\-\s]+)",
        text,
        flags=re.IGNORECASE,
    )
    if locked_match:
        locked_day = locked_match.group(1).strip().upper()[:3]
        locked_slots = _parse_slot_list(locked_match.group(2))

    return kind, pair_course, locked_day, locked_slots, "combined" in lowered


def _parse_slot_list(value: str) -> list[int]:
    slots: list[int] = []
    for chunk in re.split(r"\s*,\s*", value or ""):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "-" in chunk:
            start, end = chunk.split("-", 1)
            slots.extend(range(int(start), int(end) + 1))
        else:
            slots.append(int(chunk))
    return slots


def _is_faculty_directive(value: str) -> bool:
    lowered = value.lower().strip()
    return lowered == "auto" or lowered.startswith("x") or lowered.startswith("same-as=")


def _match_section_header(cell: str, section_lookup: dict[str, str]) -> str:
    normalized = _canonical_header(cell)
    tokens = normalized.split()
    if normalized.upper() in section_lookup:
        return section_lookup[normalized.upper()]
    if len(tokens) == 1:
        return section_lookup.get(tokens[0].upper(), "")
    for index, token in enumerate(tokens):
        candidate = token.upper()
        if candidate not in section_lookup:
            continue
        before = set(tokens[:index])
        after = set(tokens[index + 1 :])
        markers = before | after
        if markers & {"sec", "section", "faculty", "fac", "teacher", "staff"}:
            return section_lookup[candidate]
    m = re.search(r"\b(?:section|sec)\s+([a-z0-9]+)\b", normalized)
    if m:
        return section_lookup.get(m.group(1).upper(), "")
    return ""


def _canonical_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").strip().lower()).strip()


def _course_name_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def _faculty_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def _slugify(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", (value or "").upper()).strip("_")
    return slug[:32]


def _collapse_ws(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _collapse_ws(str(value))


def _clean_lines(lines: Iterable[str]) -> list[str]:
    out: list[str] = []
    for line in lines:
        cleaned = _normalize_ocr_text(line)
        if cleaned:
            out.append(cleaned)
    return out


def _normalize_ocr_text(value: str) -> str:
    text = value or ""
    text = text.replace("“", '"').replace("”", '"').replace("‘", "'").replace("’", "'")
    text = text.replace("—", "-").replace("–", "-").replace("°", "0")
    text = text.replace("ﬁ", "fi").replace("ﬂ", "fl")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _normalize_code(value: str) -> str:
    code = _collapse_ws((value or "").upper()).replace(" ", "")
    code = re.sub(r"^([A-Z]{2,})-(\d[A-Z0-9]*)$", r"\1\2", code)
    if code.startswith("8") and len(code) >= 4 and any(char.isdigit() for char in code[1:]):
        code = "B" + code[1:]
    return code


def _split_code_and_alias(code: str) -> tuple[str, str]:
    match = re.match(r"^([A-Z0-9]+)-([A-Z]{2,8})$", code)
    if match and any(ch.isdigit() for ch in match.group(1)):
        return match.group(1), match.group(2)
    return code, ""


def _parse_credit_value(value: str) -> Optional[int]:
    match = re.search(r"(?<![A-Z0-9])([0-4])(?:\s*\([^\)]*\)|\b)", value or "")
    if match:
        return int(match.group(1))
    return None


def _format_slot_list(slots: list[int]) -> str:
    cleaned = sorted({int(slot) for slot in slots})
    if not cleaned:
        return ""
    ranges: list[str] = []
    start = prev = cleaned[0]
    for slot in cleaned[1:]:
        if slot == prev + 1:
            prev = slot
            continue
        ranges.append(f"{start}-{prev}" if start != prev else str(start))
        start = prev = slot
    ranges.append(f"{start}-{prev}" if start != prev else str(start))
    return ",".join(ranges)


def _clean_candidate_name(name: str, code: str = "") -> str:
    text = _normalize_ocr_text(name)
    if code:
        compact_code = _normalize_code(code)
        text = re.sub(re.escape(compact_code), " ", text, flags=re.IGNORECASE)
        base_code, alias = _split_code_and_alias(compact_code)
        if alias:
            text = re.sub(rf"^\s*{re.escape(alias)}\b", " ", text, flags=re.IGNORECASE)
        if base_code != compact_code:
            text = re.sub(re.escape(base_code), " ", text, flags=re.IGNORECASE)
        base_match = re.match(r"^([A-Z]{2,})(\d[A-Z0-9]*)$", re.sub(r"[^A-Z0-9]", "", base_code))
        if base_match:
            letters, suffix = base_match.groups()
            text = re.sub(rf"\b{letters}\s+{suffix}\b", " ", text, flags=re.IGNORECASE)
    text = _COURSE_CODE_SPACED_RE.sub(" ", text)
    text = re.sub(rf"^({'|'.join(_COURSE_CATEGORY_PREFIXES)})\s*[-: ]\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(
        r"(?:Dr|Prof|Professor)\.?\s*[A-Za-z][A-Za-z .'-]{1,80}",
        " ",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\b\d+\(\d[^\)]*\)", " ", text)
    text = re.sub(r"\b\d+\b", " ", text)
    text = re.sub(r"\b(?:course|code|credits|faculty)\b", " ", text, flags=re.IGNORECASE)
    return _collapse_ws(text.strip(" -:|,.;"))


def _is_plausible_course_code(code: str) -> bool:
    compact = re.sub(r"[^A-Z0-9]", "", (code or "").upper())
    if len(compact) < 5:
        return False
    if sum(char.isalpha() for char in compact) < 2:
        return False
    if not any(char.isdigit() for char in compact):
        return False
    if re.fullmatch(r"(?:TR|CR)\d{2,4}[A-Z]?", compact):
        return False
    return True


def _name_quality(name: str) -> int:
    cleaned = _clean_candidate_name(name)
    lowered = cleaned.lower()
    if not cleaned:
        return 0
    if any(term in lowered for term in _TIMETABLE_NOISE_TERMS):
        return 0
    if _TITLE_RE.search(cleaned):
        return 0
    words = [word for word in re.split(r"\s+", cleaned) if word]
    alpha_chars = sum(char.isalpha() for char in cleaned)
    if alpha_chars < 3:
        return 0
    return alpha_chars + len(words) * 10


def _should_skip_free_text_line(line: str) -> bool:
    lowered = line.lower()
    if not line:
        return True
    if any(term in lowered for term in _TIMETABLE_NOISE_TERMS):
        return True
    if re.search(r"\b\d{1,2}\.\d{2}\s*-\s*\d{1,2}\.\d{2}\b", lowered):
        return True
    code_matches = list(_COURSE_CODE_RE.finditer(line.upper()))
    if len(code_matches) >= 2 and any(term in lowered for term in _DAY_OR_BREAK_TERMS):
        return True
    if lowered.startswith("4th sem ") or lowered.startswith("bms institute"):
        return True
    return False
