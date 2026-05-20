"""PDF + Excel + JSON + iCal exporters for a Timetable."""
from __future__ import annotations

import io
import json
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.units import cm

from ..models.domain import ScheduledClass, Timetable, TimetableRequest


def _columns(req: TimetableRequest) -> list[dict]:
    """Build the interleaved column descriptor list: slots + tea/lunch breaks.

    Each entry is either {"kind":"slot","slot":int} or
    {"kind":"break","label": "TEA BREAK"/"LUNCH BREAK", "duration": int}.
    """
    tc = req.time_config
    cols: list[dict] = []
    for t in range(1, tc.slots_per_day + 1):
        cols.append({"kind": "slot", "slot": t})
        if t == tc.tea_break.after_slot:
            cols.append(
                {"kind": "break", "label": "TEA BREAK", "duration": tc.tea_break.duration_min}
            )
        if t == tc.lunch_break.after_slot:
            cols.append(
                {"kind": "break", "label": "LUNCH BREAK", "duration": tc.lunch_break.duration_min}
            )
    return cols


def _slot_timing(req: TimetableRequest, t: int) -> str:
    timings = req.time_config.slot_timings
    if t - 1 < len(timings):
        return f"{timings[t - 1].start}–{timings[t - 1].end}"
    return ""


def _cell_paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(text).replace("\n", "<br/>"), style)


def _clean_faculty_name(name: str) -> str:
    return name.split(" (", 1)[0].strip()


def _dot_time_12h(hhmm: str) -> str:
    try:
        hh, mm = [int(x) for x in hhmm.split(":")]
        hh = hh % 12 or 12
        return f"{hh}.{mm:02d}"
    except Exception:
        return hhmm.replace(":", ".")


def _slot_timing_template(req: TimetableRequest, t: int) -> str:
    timings = req.time_config.slot_timings
    if t - 1 >= len(timings):
        return ""
    tm = timings[t - 1]
    return f"{_dot_time_12h(tm.start)}-{_dot_time_12h(tm.end)}"


def _break_timing_template(req: TimetableRequest, after_slot: int, duration_min: int) -> str:
    timings = req.time_config.slot_timings
    if after_slot - 1 >= len(timings):
        return ""
    try:
        end_str = timings[after_slot - 1].end
        start_dt = datetime.strptime(end_str, "%H:%M")
        end_dt = start_dt + timedelta(minutes=duration_min)
        return f"{_dot_time_12h(start_dt.strftime('%H:%M'))}-{_dot_time_12h(end_dt.strftime('%H:%M'))}"
    except Exception:
        return ""


def _roman(n: int) -> str:
    vals = {
        1: "I",
        2: "II",
        3: "III",
        4: "IV",
        5: "V",
        6: "VI",
        7: "VII",
        8: "VIII",
        9: "IX",
        10: "X",
    }
    return vals.get(n, str(n))


def _full_day_label(day: str) -> str:
    return {
        "MON": "MONDAY",
        "TUE": "TUESDAY",
        "WED": "WEDNESDAY",
        "THU": "THURSDAY",
        "FRI": "FRIDAY",
        "SAT": "SATURDAY",
        "SUN": "SUNDAY",
    }.get(day, day)


def _saturday_note(req: TimetableRequest) -> str:
    weeks = req.time_config.saturday_rules.inactive_weeks
    if not weeks:
        return "Note: Saturday classes as per timetable."
    ord_map = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}
    parts = [ord_map.get(w, f"{w}th") for w in weeks]
    if len(parts) == 1:
        txt = parts[0]
    else:
        txt = ", ".join(parts[:-1]) + " and " + parts[-1]
    return f"Note: No Classes on {txt} Saturday of every Month"


# ---------------------------------------------------------------------------
# Grid building
# ---------------------------------------------------------------------------
def _section_grid(tt: Timetable, sec_id: str, days: list[str], slots: list[int]):
    grid: dict[tuple[str, int], list[str]] = defaultdict(list)
    for c in tt.classes:
        if c.section_id != sec_id:
            continue
        cell = grid[(c.day, c.slot)]
        label = c.label or c.course_code
        batch = f" ({c.batch_id})" if c.batch_id else ""
        full = f"{label}{batch}"
        if full not in cell:
            cell.append(full)
    return grid


def _faculty_grid(tt: Timetable, faculty_id: str, days: list[str], slots: list[int]):
    grid: dict[tuple[str, int], list[str]] = defaultdict(list)
    for c in tt.classes:
        if c.faculty_id != faculty_id:
            continue
        cell = grid[(c.day, c.slot)]
        text = f"{c.section_id}: {c.label or c.course_code}"
        if text not in cell:
            cell.append(text)
    return grid


def _faculty_classes(req: TimetableRequest, tt: Timetable, faculty_ids: list[str]) -> list[ScheduledClass]:
    id_set = {fid.strip() for fid in faculty_ids if fid and fid.strip()}
    if not id_set:
        return []

    faculty_by_id = {fac.id: fac for fac in req.faculty}
    selected = [faculty_by_id[fid] for fid in faculty_ids if fid in faculty_by_id]
    out: list[ScheduledClass] = []
    seen: set[str] = set()
    day_order = {day: idx for idx, day in enumerate(req.time_config.days)}

    def push_unique(scheduled: ScheduledClass) -> None:
        key = "|".join(
            [
                scheduled.section_id,
                scheduled.day,
                str(scheduled.slot),
                scheduled.course_code,
                scheduled.label or "",
                scheduled.batch_id or "",
            ]
        )
        if key in seen:
            return
        seen.add(key)
        out.append(scheduled)

    for scheduled in tt.classes:
        if scheduled.faculty_id and scheduled.faculty_id in id_set:
            push_unique(scheduled)

    for fac in selected:
        for assignment in fac.assignments:
            for scheduled in tt.classes:
                if (
                    scheduled.course_code == assignment.course_code
                    and scheduled.section_id == assignment.section_id
                ):
                    push_unique(scheduled)

    for block in req.elective_blocks:
        for opt in block.options:
            if not any(fid in id_set for fid in opt.faculty_pool):
                continue
            for scheduled in tt.classes:
                if scheduled.course_code != block.id:
                    continue
                if block.applies_to_sections and scheduled.section_id not in block.applies_to_sections:
                    continue
                push_unique(
                    scheduled.model_copy(
                        update={
                            "course_code": opt.course_code or scheduled.course_code,
                            "label": opt.course_name or scheduled.label or scheduled.course_code,
                        }
                    )
                )

    lower_names = [fac.name.lower() for fac in selected]
    wants_iic = any("iic" in name for name in lower_names)
    wants_beng = any("bengdip2" in name for name in lower_names)
    if wants_iic or wants_beng:
        for scheduled in tt.classes:
            code = (scheduled.course_code or "").lower()
            label = (scheduled.label or "").lower()
            if (
                wants_iic
                and ("iic" in code or "iic" in label)
            ) or (
                wants_beng
                and ("bengdip2" in code or "bengdip2" in label)
            ):
                push_unique(scheduled)

    out.sort(
        key=lambda scheduled: (
            day_order.get(scheduled.day, 99),
            scheduled.slot,
            scheduled.section_id,
            scheduled.course_code,
            scheduled.batch_id or "",
        )
    )
    return out


def _faculty_grid_for_classes(classes: list[ScheduledClass]):
    grid: dict[tuple[str, int], list[str]] = defaultdict(list)
    for scheduled in classes:
        cell = grid[(scheduled.day, scheduled.slot)]
        batch = f" ({scheduled.batch_id})" if scheduled.batch_id else ""
        text = f"{scheduled.section_id}: {scheduled.label or scheduled.course_code}{batch}"
        if text not in cell:
            cell.append(text)
    return grid


def _faculty_course_rows(req: TimetableRequest, classes: list[ScheduledClass]) -> list[dict[str, str]]:
    course_by_code = {course.code: course for course in req.courses}
    rows: dict[str, dict[str, object]] = {}
    for scheduled in classes:
        code = scheduled.course_code
        course = course_by_code.get(code)
        if code not in rows:
            rows[code] = {
                "course_name": course.name if course else (scheduled.label or code),
                "course_code": code,
                "sections": set(),
                "order": len(rows),
            }
        rows[code]["sections"].add(scheduled.section_id)

    out: list[dict[str, str]] = []
    for row in rows.values():
        sections = ", ".join(sorted(row["sections"]))
        out.append(
            {
                "course_name": str(row["course_name"]),
                "course_code": str(row["course_code"]),
                "sections": sections or "—",
            }
        )
    out.sort(key=lambda row: (row["course_name"], row["course_code"]))
    return out


def _section_faculty_rows(req: TimetableRequest, tt: Timetable, sec_id: str) -> list[dict[str, str]]:
    course_by_code = {c.code: c for c in req.courses}
    faculty_by_id = {f.id: _clean_faculty_name(f.name) for f in req.faculty}
    block_by_id = {b.id: b for b in req.elective_blocks}
    course_order = {c.code: i for i, c in enumerate(req.courses)}
    rows: dict[str, dict] = {}

    def upsert(course_code: str, fallback_name: str | None = None) -> dict:
        course = course_by_code.get(course_code)
        course_name = course.name if course else (fallback_name or course_code)
        credits = str(course.credits) if course else "—"
        if course_code not in rows:
            rows[course_code] = {
                "course_name": course_name,
                "course_code": course_code,
                "credits": credits,
                "faculty": set(),
                "order": course_order.get(course_code, 10_000 + len(rows)),
            }
        return rows[course_code]

    for fac in req.faculty:
        for assign in fac.assignments:
            if assign.section_id != sec_id:
                continue
            entry = upsert(assign.course_code)
            entry["faculty"].add(fac.name)

    section_classes = [c for c in tt.classes if c.section_id == sec_id]
    elective_block_ids = {
        c.course_code for c in section_classes if c.course_code.startswith("ELEC_BLOCK_")
    }
    for block_id in elective_block_ids:
        block = block_by_id.get(block_id)
        if not block:
            continue
        for opt in block.options:
            entry = upsert(opt.course_code, opt.course_name)
            for fid in opt.faculty_pool:
                name = faculty_by_id.get(fid)
                if name:
                    entry["faculty"].add(name)

    for c in section_classes:
        if c.course_code.startswith("ELEC_BLOCK_"):
            continue
        entry = upsert(c.course_code, c.label)
        if c.faculty_id:
            entry["faculty"].add(faculty_by_id.get(c.faculty_id, c.faculty_id))

    out: list[dict[str, str]] = []
    for row in sorted(rows.values(), key=lambda r: (r["order"], r["course_name"])):
        faculty_names = ", ".join(sorted(row["faculty"])) if row["faculty"] else "—"
        out.append(
            {
                "course_name": row["course_name"],
                "course_code": row["course_code"],
                "credits": row["credits"],
                "faculty": faculty_names,
            }
        )
    return out


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def _render_section_pdf(req: TimetableRequest, tt: Timetable) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        title="Timetable",
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )
    styles = getSampleStyleSheet()
    elements: list = []

    days = req.time_config.days
    slots = list(range(1, req.time_config.slots_per_day + 1))
    cols = _columns(req)
    break_col_idxs = [i + 1 for i, c in enumerate(cols) if c["kind"] == "break"]
    slot_count = sum(1 for c in cols if c["kind"] == "slot")

    page_w = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    break_w = 1.9 * cm
    day_w = 2.8 * cm
    remaining = page_w - day_w - break_w * len(break_col_idxs)
    slot_w = max(remaining / max(slot_count, 1), 1.8 * cm)
    col_widths = [day_w]
    for c in cols:
        col_widths.append(break_w if c["kind"] == "break" else slot_w)

    institute_style = ParagraphStyle(
        "InstituteHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=15,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    dept_style = ParagraphStyle(
        "DeptHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=10,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    meta_style = ParagraphStyle(
        "MetaHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    grid_header_style = ParagraphStyle(
        "GridHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    time_style = ParagraphStyle(
        "GridTime",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    day_style = ParagraphStyle(
        "GridDay",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    body_style = ParagraphStyle(
        "GridBody",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.6,
        leading=8.6,
        alignment=1,
        wordWrap="CJK",
        spaceBefore=0,
        spaceAfter=0,
    )
    break_style = ParagraphStyle(
        "GridBreak",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=9.2,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    note_style = ParagraphStyle(
        "Note",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    faculty_header_style = ParagraphStyle(
        "FacultyHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    faculty_body_style = ParagraphStyle(
        "FacultyBody",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.4,
        leading=9,
        alignment=1,
        wordWrap="CJK",
        spaceBefore=0,
        spaceAfter=0,
    )
    tea_range = _break_timing_template(
        req, req.time_config.tea_break.after_slot, req.time_config.tea_break.duration_min
    )
    lunch_range = _break_timing_template(
        req, req.time_config.lunch_break.after_slot, req.time_config.lunch_break.duration_min
    )

    for sec_idx, sec in enumerate(req.sections):
        effective_from = datetime.now()
        top_data = [
            [_cell_paragraph("BMS INSTITUTE OF TECHNOLOGY AND MANAGEMENT", institute_style), "", ""],
            [
                _cell_paragraph(
                    "DEPARTMENT OF ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING",
                    dept_style,
                ),
                "",
                "",
            ],
            [_cell_paragraph("Academic Year: 2025-26 (EVEN Sem)", meta_style), "", ""],
            [
                _cell_paragraph(sec.name, meta_style),
                _cell_paragraph("Class Time Table", meta_style),
                _cell_paragraph(f"Class Room: {sec.classroom or '-'}", meta_style),
            ],
            [
                _cell_paragraph(
                    f"With Effect From: {effective_from.day}/{effective_from.month}/{effective_from.year}",
                    meta_style,
                ),
                "",
                _cell_paragraph("Version:01", meta_style),
            ],
        ]
        top_tbl = Table(top_data, colWidths=[page_w * 0.28, page_w * 0.44, page_w * 0.28], hAlign="LEFT")
        top_tbl.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                    ("BOX", (0, 0), (-1, -1), 1.1, colors.black),
                    ("SPAN", (0, 0), (-1, 0)),
                    ("SPAN", (0, 1), (-1, 1)),
                    ("SPAN", (0, 2), (-1, 2)),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        elements.append(top_tbl)

        grid = _section_grid(tt, sec.id, days, slots)
        raw_data: list[list[str]] = []
        roman_row = [""]
        time_row = [""]
        for col in cols:
            if col["kind"] == "slot":
                roman_row.append(_roman(col["slot"]))
                time_row.append(_slot_timing_template(req, col["slot"]))
            else:
                roman_row.append("")
                if col["label"] == "TEA BREAK":
                    time_row.append(tea_range)
                elif col["label"] == "LUNCH BREAK":
                    time_row.append(lunch_range)
                else:
                    time_row.append("")
        raw_data.append(roman_row)
        raw_data.append(time_row)

        for d in days:
            row = [_full_day_label(d)]
            for col in cols:
                if col["kind"] == "break":
                    row.append("")
                    continue
                cell = grid.get((d, col["slot"]))
                row.append(" / ".join(cell) if cell else "")
            raw_data.append(row)

        first_day_row = 2
        for bc in break_col_idxs:
            raw_data[first_day_row][bc] = cols[bc - 1]["label"].replace(" ", "\n")
            for r in range(first_day_row + 1, len(raw_data)):
                raw_data[r][bc] = ""

        merged_cells: list[tuple[int, int, int]] = []
        for r in range(first_day_row, len(raw_data)):
            cidx = 1
            while cidx < len(raw_data[r]):
                if cidx in break_col_idxs:
                    cidx += 1
                    continue
                txt = raw_data[r][cidx].strip()
                if not txt:
                    cidx += 1
                    continue
                end = cidx
                while (
                    end + 1 < len(raw_data[r])
                    and (end + 1) not in break_col_idxs
                    and raw_data[r][end + 1].strip() == txt
                ):
                    end += 1
                if end > cidx:
                    merged_cells.append((r, cidx, end))
                cidx = end + 1

        data: list[list] = []
        for r, row in enumerate(raw_data):
            out_row: list = []
            for cidx, txt in enumerate(row):
                if not txt:
                    out_row.append("")
                    continue
                if r == 0:
                    out_row.append(_cell_paragraph(txt, grid_header_style))
                elif r == 1:
                    out_row.append(_cell_paragraph(txt, time_style))
                elif cidx == 0:
                    out_row.append(_cell_paragraph(txt, day_style))
                elif cidx in break_col_idxs:
                    out_row.append(_cell_paragraph(txt, break_style))
                else:
                    out_row.append(_cell_paragraph(txt, body_style))
            data.append(out_row)

        row_heights = [0.50 * cm, 0.60 * cm] + [0.92 * cm for _ in days]
        tbl = Table(data, colWidths=col_widths, rowHeights=row_heights, hAlign="LEFT")
        style_cmds: list = [
            ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
            ("BOX", (0, 0), (-1, -1), 1.1, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, 1), "CENTER"),
            ("ALIGN", (0, 2), (0, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTNAME", (0, 2), (0, -1), "Helvetica-Bold"),
        ]
        for bc in break_col_idxs:
            style_cmds.append(("SPAN", (bc, first_day_row), (bc, len(data) - 1)))
            style_cmds.append(("BACKGROUND", (bc, 0), (bc, -1), colors.HexColor("#F5F5F5")))
            style_cmds.append(("ALIGN", (bc, 0), (bc, -1), "CENTER"))
        for r, start, end in merged_cells:
            style_cmds.append(("SPAN", (start, r), (end, r)))
        tbl.setStyle(TableStyle(style_cmds))
        elements.append(tbl)

        note_tbl = Table(
            [[_cell_paragraph(_saturday_note(req), note_style)]],
            colWidths=[page_w],
            hAlign="LEFT",
        )
        note_tbl.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                    ("BOX", (0, 0), (-1, -1), 1.1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        elements.append(note_tbl)

        faculty_rows = _section_faculty_rows(req, tt, sec.id)
        if faculty_rows:
            split_idx = (len(faculty_rows) + 1) // 2
            left_rows = faculty_rows[:split_idx]
            right_rows = faculty_rows[split_idx:]
            n_rows = max(len(left_rows), len(right_rows))
            faculty_data: list[list] = [
                [
                    _cell_paragraph("COURSE NAME", faculty_header_style),
                    _cell_paragraph("COURSE CODE", faculty_header_style),
                    _cell_paragraph("CREDITS", faculty_header_style),
                    _cell_paragraph("FACULTY", faculty_header_style),
                    _cell_paragraph("COURSE NAME", faculty_header_style),
                    _cell_paragraph("COURSE CODE", faculty_header_style),
                    _cell_paragraph("CREDITS", faculty_header_style),
                    _cell_paragraph("FACULTY", faculty_header_style),
                ]
            ]
            for i in range(n_rows):
                lrow = left_rows[i] if i < len(left_rows) else None
                rrow = right_rows[i] if i < len(right_rows) else None
                faculty_data.append(
                    [
                        _cell_paragraph(lrow["course_name"] if lrow else "", faculty_body_style),
                        _cell_paragraph(lrow["course_code"] if lrow else "", faculty_body_style),
                        _cell_paragraph(lrow["credits"] if lrow else "", faculty_body_style),
                        _cell_paragraph(lrow["faculty"] if lrow else "", faculty_body_style),
                        _cell_paragraph(rrow["course_name"] if rrow else "", faculty_body_style),
                        _cell_paragraph(rrow["course_code"] if rrow else "", faculty_body_style),
                        _cell_paragraph(rrow["credits"] if rrow else "", faculty_body_style),
                        _cell_paragraph(rrow["faculty"] if rrow else "", faculty_body_style),
                    ]
                )

            half_w = page_w / 2
            cname_w = 3.5 * cm
            ccode_w = 2.2 * cm
            cred_w = 1.3 * cm
            fac_w = half_w - cname_w - ccode_w - cred_w
            faculty_col_widths = [cname_w, ccode_w, cred_w, fac_w, cname_w, ccode_w, cred_w, fac_w]
            fac_tbl = Table(faculty_data, colWidths=faculty_col_widths, hAlign="LEFT")
            fac_tbl.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                        ("BOX", (0, 0), (-1, -1), 1.1, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                        ("ALIGN", (1, 1), (2, -1), "CENTER"),
                        ("ALIGN", (5, 1), (6, -1), "CENTER"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 2),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ]
                )
            )
            elements.append(fac_tbl)

        if sec_idx < len(req.sections) - 1:
            elements.append(PageBreak())

    if not req.sections:
        elements.append(Paragraph("No data", styles["Normal"]))
    doc.build(elements)
    return buf.getvalue()


def _render_faculty_pdf(req: TimetableRequest, tt: Timetable, faculty_ids: list[str]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        title="Faculty Timetable",
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )
    styles = getSampleStyleSheet()
    elements: list = []

    days = req.time_config.days
    slots = list(range(1, req.time_config.slots_per_day + 1))
    cols = _columns(req)
    break_col_idxs = [i + 1 for i, c in enumerate(cols) if c["kind"] == "break"]
    slot_count = sum(1 for c in cols if c["kind"] == "slot")

    page_w = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    break_w = 1.9 * cm
    day_w = 2.8 * cm
    remaining = page_w - day_w - break_w * len(break_col_idxs)
    slot_w = max(remaining / max(slot_count, 1), 1.8 * cm)
    col_widths = [day_w]
    for col in cols:
        col_widths.append(break_w if col["kind"] == "break" else slot_w)

    institute_style = ParagraphStyle(
        "InstituteHeaderFaculty",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=15,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    dept_style = ParagraphStyle(
        "DeptHeaderFaculty",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=10,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    meta_style = ParagraphStyle(
        "MetaHeaderFaculty",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    grid_header_style = ParagraphStyle(
        "FacultyGridHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    time_style = ParagraphStyle(
        "FacultyGridTime",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    day_style = ParagraphStyle(
        "FacultyGridDay",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    body_style = ParagraphStyle(
        "FacultyGridBody",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.6,
        leading=8.6,
        alignment=1,
        wordWrap="CJK",
        spaceBefore=0,
        spaceAfter=0,
    )
    break_style = ParagraphStyle(
        "FacultyGridBreak",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=9.2,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    note_style = ParagraphStyle(
        "FacultyNote",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    summary_header_style = ParagraphStyle(
        "FacultySummaryHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=1,
        spaceBefore=0,
        spaceAfter=0,
    )
    summary_body_style = ParagraphStyle(
        "FacultySummaryBody",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.4,
        leading=9,
        alignment=1,
        wordWrap="CJK",
        spaceBefore=0,
        spaceAfter=0,
    )

    faculty_classes = _faculty_classes(req, tt, faculty_ids)
    faculty_names = [
        _clean_faculty_name(fac.name)
        for fac in req.faculty
        if fac.id in {fid.strip() for fid in faculty_ids if fid.strip()}
    ]
    faculty_names = list(dict.fromkeys(faculty_names))
    faculty_label = ", ".join(faculty_names) if faculty_names else ", ".join(faculty_ids)
    course_rows = _faculty_course_rows(req, faculty_classes)
    effective_from = datetime.now()
    tea_range = _break_timing_template(
        req, req.time_config.tea_break.after_slot, req.time_config.tea_break.duration_min
    )
    lunch_range = _break_timing_template(
        req, req.time_config.lunch_break.after_slot, req.time_config.lunch_break.duration_min
    )

    top_data = [
        [_cell_paragraph("BMS INSTITUTE OF TECHNOLOGY AND MANAGEMENT", institute_style), "", ""],
        [
            _cell_paragraph(
                "DEPARTMENT OF ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING",
                dept_style,
            ),
            "",
            "",
        ],
        [_cell_paragraph("Academic Year: 2025-26 (EVEN Sem)", meta_style), "", ""],
        [
            _cell_paragraph(faculty_label or "Faculty", meta_style),
            _cell_paragraph("Faculty Time Table", meta_style),
            _cell_paragraph(f"Courses: {len(course_rows)}", meta_style),
        ],
        [
            _cell_paragraph(
                f"With Effect From: {effective_from.day}/{effective_from.month}/{effective_from.year}",
                meta_style,
            ),
            "",
            _cell_paragraph(f"Weekly Slots: {len(faculty_classes)}", meta_style),
        ],
    ]
    top_tbl = Table(top_data, colWidths=[page_w * 0.28, page_w * 0.44, page_w * 0.28], hAlign="LEFT")
    top_tbl.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("BOX", (0, 0), (-1, -1), 1.1, colors.black),
                ("SPAN", (0, 0), (-1, 0)),
                ("SPAN", (0, 1), (-1, 1)),
                ("SPAN", (0, 2), (-1, 2)),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    elements.append(top_tbl)

    grid = _faculty_grid_for_classes(faculty_classes)
    raw_data: list[list[str]] = []
    roman_row = [""]
    time_row = [""]
    for col in cols:
        if col["kind"] == "slot":
            roman_row.append(_roman(col["slot"]))
            time_row.append(_slot_timing_template(req, col["slot"]))
        else:
            roman_row.append("")
            if col["label"] == "TEA BREAK":
                time_row.append(tea_range)
            elif col["label"] == "LUNCH BREAK":
                time_row.append(lunch_range)
            else:
                time_row.append("")
    raw_data.append(roman_row)
    raw_data.append(time_row)

    for day in days:
        row = [_full_day_label(day)]
        for col in cols:
            if col["kind"] == "break":
                row.append("")
                continue
            cell = grid.get((day, col["slot"]))
            row.append("\n".join(cell) if cell else "")
        raw_data.append(row)

    first_day_row = 2
    for bc in break_col_idxs:
        raw_data[first_day_row][bc] = cols[bc - 1]["label"].replace(" ", "\n")
        for r in range(first_day_row + 1, len(raw_data)):
            raw_data[r][bc] = ""

    merged_cells: list[tuple[int, int, int]] = []
    for r in range(first_day_row, len(raw_data)):
        cidx = 1
        while cidx < len(raw_data[r]):
            if cidx in break_col_idxs:
                cidx += 1
                continue
            txt = raw_data[r][cidx].strip()
            if not txt:
                cidx += 1
                continue
            end = cidx
            while (
                end + 1 < len(raw_data[r])
                and (end + 1) not in break_col_idxs
                and raw_data[r][end + 1].strip() == txt
            ):
                end += 1
            if end > cidx:
                merged_cells.append((r, cidx, end))
            cidx = end + 1

    data: list[list] = []
    for r, row in enumerate(raw_data):
        out_row: list = []
        for cidx, txt in enumerate(row):
            if not txt:
                out_row.append("")
                continue
            if r == 0:
                out_row.append(_cell_paragraph(txt, grid_header_style))
            elif r == 1:
                out_row.append(_cell_paragraph(txt, time_style))
            elif cidx == 0:
                out_row.append(_cell_paragraph(txt, day_style))
            elif cidx in break_col_idxs:
                out_row.append(_cell_paragraph(txt, break_style))
            else:
                out_row.append(_cell_paragraph(txt, body_style))
        data.append(out_row)

    row_heights = [0.50 * cm, 0.60 * cm] + [0.92 * cm for _ in days]
    tbl = Table(data, colWidths=col_widths, rowHeights=row_heights, hAlign="LEFT")
    style_cmds: list = [
        ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
        ("BOX", (0, 0), (-1, -1), 1.1, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 1), "CENTER"),
        ("ALIGN", (0, 2), (0, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTNAME", (0, 2), (0, -1), "Helvetica-Bold"),
    ]
    for bc in break_col_idxs:
        style_cmds.append(("SPAN", (bc, first_day_row), (bc, len(data) - 1)))
        style_cmds.append(("BACKGROUND", (bc, 0), (bc, -1), colors.HexColor("#F5F5F5")))
        style_cmds.append(("ALIGN", (bc, 0), (bc, -1), "CENTER"))
    for r, start, end in merged_cells:
        style_cmds.append(("SPAN", (start, r), (end, r)))
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    note_text = (
        _saturday_note(req)
        if faculty_classes
        else "No scheduled classes found for the selected faculty."
    )
    note_tbl = Table(
        [[_cell_paragraph(note_text, note_style)]],
        colWidths=[page_w],
        hAlign="LEFT",
    )
    note_tbl.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("BOX", (0, 0), (-1, -1), 1.1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    elements.append(note_tbl)

    if course_rows:
        summary_data: list[list] = [
            [
                _cell_paragraph("COURSE NAME", summary_header_style),
                _cell_paragraph("COURSE CODE", summary_header_style),
                _cell_paragraph("SECTIONS", summary_header_style),
            ]
        ]
        for row in course_rows:
            summary_data.append(
                [
                    _cell_paragraph(row["course_name"], summary_body_style),
                    _cell_paragraph(row["course_code"], summary_body_style),
                    _cell_paragraph(row["sections"], summary_body_style),
                ]
            )
        summary_tbl = Table(
            summary_data,
            colWidths=[page_w * 0.5, page_w * 0.2, page_w * 0.3],
            hAlign="LEFT",
        )
        summary_tbl.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                    ("BOX", (0, 0), (-1, -1), 1.1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("ALIGN", (1, 1), (2, -1), "CENTER"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        elements.append(summary_tbl)

    doc.build(elements)
    return buf.getvalue()


def render_pdf(req: TimetableRequest, tt: Timetable, faculty_ids: list[str] | None = None) -> bytes:
    if faculty_ids:
        return _render_faculty_pdf(req, tt, faculty_ids)
    return _render_section_pdf(req, tt)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def _excel_logo_path() -> Path | None:
    repo_root = Path(__file__).resolve().parents[3]
    candidates = [
        repo_root / "frontend" / "public" / "bmsitm-logo.png",
        repo_root / "frontend" / "dist" / "bmsitm-logo.png",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _attach_excel_logo(ws, anchor: str = "A1") -> None:
    logo_path = _excel_logo_path()
    if not logo_path:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        return
    try:
        image = XLImage(str(logo_path))
        image.width = 64
        image.height = 64
        ws.add_image(image, anchor)
    except Exception:
        return


def _apply_excel_heading(
    ws,
    *,
    last_column: int,
    title: str,
    subtitle: str,
    meta: str,
) -> int:
    title_fill = PatternFill("solid", fgColor="0E2447")
    subtitle_fill = PatternFill("solid", fgColor="173561")
    meta_fill = PatternFill("solid", fgColor="F7E8BF")
    title_font = Font(bold=True, color="FFFFFF", size=15)
    subtitle_font = Font(bold=True, color="FFFFFF", size=10)
    meta_font = Font(bold=True, color="0E2447", size=9)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 8

    ws.column_dimensions["A"].width = 11
    _attach_excel_logo(ws, "A1")

    heading_rows = [
        ("BMS INSTITUTE OF TECHNOLOGY AND MANAGEMENT", title_fill, title_font),
        ("DEPARTMENT OF ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING", subtitle_fill, subtitle_font),
        ("Academic Year: 2025-26 (EVEN Sem)", subtitle_fill, subtitle_font),
        (title, meta_fill, meta_font),
        (f"{subtitle} | {meta}", meta_fill, meta_font),
    ]

    for row_index, (text, fill, font) in enumerate(heading_rows, start=1):
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=last_column)
        cell = ws.cell(row=row_index, column=2, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = header_align

    return 7


def render_xlsx(req: TimetableRequest, tt: Timetable) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    days = req.time_config.days
    slots = list(range(1, req.time_config.slots_per_day + 1))
    cols = _columns(req)

    header_fill = PatternFill("solid", fgColor="0E2447")
    header_font = Font(bold=True, color="FFFFFF")
    day_fill = PatternFill("solid", fgColor="F0E5C9")
    day_font = Font(bold=True, color="0E2447")
    break_fill = PatternFill("solid", fgColor="F4C752")
    break_font = Font(bold=True, color="161412", size=11)
    body_font = Font(size=9)
    body_alt = PatternFill("solid", fgColor="FFFBED")
    today_label = datetime.now().strftime("%d/%m/%Y")

    for sec in req.sections:
        ws = wb.create_sheet(title=f"Sec {sec.id}")
        last_column = len(cols) + 1
        table_header_row = _apply_excel_heading(
            ws,
            last_column=last_column,
            title=f"{sec.name} | Class Time Table",
            subtitle=f"Class Room: {sec.classroom or '-'}",
            meta=f"With Effect From: {today_label} | Version: 01",
        )
        first_data_row = table_header_row + 1
        # Column 1 = Day, columns 2.. = slots interleaved with breaks
        ws.cell(row=table_header_row, column=1, value="Day")
        for j, c in enumerate(cols, start=2):
            if c["kind"] == "slot":
                t = c["slot"]
                ws.cell(row=table_header_row, column=j, value=f"Slot {t}\n{_slot_timing(req, t)}")
            else:
                ws.cell(row=table_header_row, column=j, value=c["label"])
        for cell in ws[table_header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[table_header_row].height = 32

        grid = _section_grid(tt, sec.id, days, slots)
        for i, d in enumerate(days, start=first_data_row):
            day_cell = ws.cell(row=i, column=1, value=d)
            day_cell.fill = day_fill
            day_cell.font = day_font
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            for j, c in enumerate(cols, start=2):
                if c["kind"] == "break":
                    ws.cell(row=i, column=j, value="")
                    continue
                t = c["slot"]
                cell_items = grid.get((d, t))
                txt = "\n".join(cell_items) if cell_items else ("—" if d == "SAT" else "")
                cc = ws.cell(row=i, column=j, value=txt)
                cc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                cc.font = body_font
                if i % 2 == 0:
                    cc.fill = body_alt
            ws.row_dimensions[i].height = 46

        # Merge break columns vertically (header row through last data row) so
        # the label sits centred over its column, mirroring the UI grid.
        last_row = table_header_row + len(days)
        for j, c in enumerate(cols, start=2):
            if c["kind"] != "break":
                continue
            col_letter = ws.cell(row=table_header_row, column=j).column_letter
            ws.merge_cells(start_row=table_header_row, start_column=j, end_row=last_row, end_column=j)
            merged = ws.cell(row=table_header_row, column=j)
            merged.value = c["label"].replace(" ", "\n")
            merged.fill = break_fill
            merged.font = break_font
            merged.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True, text_rotation=90
            )
            ws.column_dimensions[col_letter].width = 5

        # Column widths
        for col in range(1, len(cols) + 2):
            if col == 1:
                ws.column_dimensions[ws.cell(row=table_header_row, column=col).column_letter].width = 11
            else:
                desc = cols[col - 2]
                if desc["kind"] == "slot":
                    ws.column_dimensions[ws.cell(row=table_header_row, column=col).column_letter].width = 20
        ws.freeze_panes = f"B{first_data_row}"

    # Faculty sheet
    fws = wb.create_sheet(title="Faculty view")
    faculty_header_row = _apply_excel_heading(
        fws,
        last_column=5,
        title="Faculty Time Table",
        subtitle="Section-wise faculty schedule",
        meta=f"With Effect From: {today_label}",
    )
    faculty_columns = ["Faculty", "Day", "Slot", "Section", "Course"]
    for column_index, value in enumerate(faculty_columns, start=1):
        fws.cell(row=faculty_header_row, column=column_index, value=value)
    for cell in fws[faculty_header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fws.row_dimensions[faculty_header_row].height = 26
    faculty_data_row = faculty_header_row + 1
    for c in tt.classes:
        if c.faculty_id:
            fws.cell(row=faculty_data_row, column=1, value=c.faculty_id)
            fws.cell(row=faculty_data_row, column=2, value=c.day)
            fws.cell(row=faculty_data_row, column=3, value=c.slot)
            fws.cell(row=faculty_data_row, column=4, value=c.section_id)
            fws.cell(row=faculty_data_row, column=5, value=c.label or c.course_code)
            faculty_data_row += 1
    for column in range(1, 6):
        width = 18 if column in (1, 5) else 12
        fws.column_dimensions[fws.cell(row=faculty_header_row, column=column).column_letter].width = width
    fws.freeze_panes = f"A{faculty_header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------
def render_json(req: TimetableRequest, tt: Timetable) -> bytes:
    return json.dumps(
        {
            "request": req.model_dump(),
            "timetable": tt.model_dump(),
        },
        default=str,
        indent=2,
    ).encode("utf-8")


# ---------------------------------------------------------------------------
# iCal (per faculty)
# ---------------------------------------------------------------------------
def render_ical(req: TimetableRequest, tt: Timetable, faculty_id: str) -> bytes:
    from icalendar import Calendar, Event

    cal = Calendar()
    cal.add("prodid", "-//Timetable Generator//EN")
    cal.add("version", "2.0")

    # Use next Monday as week start
    today = datetime.now().date()
    monday = today + timedelta(days=(7 - today.weekday()) % 7)
    day_offset = {d: i for i, d in enumerate(["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"])}

    for c in tt.classes:
        if c.faculty_id != faculty_id:
            continue
        timing = (
            req.time_config.slot_timings[c.slot - 1]
            if c.slot - 1 < len(req.time_config.slot_timings)
            else None
        )
        if not timing:
            continue
        d = monday + timedelta(days=day_offset.get(c.day, 0))
        start_h, start_m = [int(x) for x in timing.start.split(":")]
        end_h, end_m = [int(x) for x in timing.end.split(":")]
        ev = Event()
        ev.add("summary", f"{c.label or c.course_code} ({c.section_id})")
        ev.add("dtstart", datetime(d.year, d.month, d.day, start_h, start_m))
        ev.add("dtend", datetime(d.year, d.month, d.day, end_h, end_m))
        ev.add("description", f"Section {c.section_id} | {c.course_code}")
        cal.add_component(ev)

    return bytes(cal.to_ical())
